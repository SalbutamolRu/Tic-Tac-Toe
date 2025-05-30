#importing library
import turtle
import openpyxl
import time
import random
from openpyxl import Workbook,load_workbook

def scoreLogin():
	global player_one
	global player_two
	global playing_char
	global player_char
	global steps_one
	global steps_two
	global win
	global draw
	global mode
	
	played = False
	file = load_workbook('ScoreTable.xlsx')
	table = file.active
	maxRow = table.max_row
	playerRowOne = 0
	playerRowTwo = 0

	if mode == 1:
		for i in range(2, maxRow+1):
				temp = table.cell(row = i, column = 1).value
				if  temp == player_one:
					table.cell(row = i, column = 3).value += 1
					try:
						table.cell(row = i, column = 5).value += 1
					except:
						table.cell(row = i,column = 5).value = 1
					try:
						table.cell(row = i, column = 6).value += 1
					except:
						table.cell(row = i, column = 6).value = 1
					table.cell(row = i, column = 9).value = (steps_one + table.cell(row = i, column = 9).value)/2
					playerRowOne = i
				elif temp == player_two:
					table.cell(row = i, column = 3).value += 1
					try:
						table.cell(row = i, column = 4).value += 1
					except:
						table.cell(row = i,column = 4).value = 1
					try:
						table.cell(row = i, column = 6).value += 1
					except:
						table.cell(row = i, column = 6).value = 1
					table.cell(row = i, column = 9).value = (steps_one + table.cell(row = i, column = 9).value)/2
					playerRowTwo = i
		if playerRowOne == 0:
			maxRow += 1
			playerRowOne = maxRow
			table.cell(row = playerRowOne, column = 1).value = player_one
			table.cell(row = playerRowOne, column = 11).value = 'E'
			table.cell(row = playerRowOne, column = 3).value = 1
			table.cell(row = playerRowOne, column = 5).value = 1
			table.cell(row = playerRowOne, column = 6).value = 1
			table.cell(row = playerRowOne, column = 9).value = steps_one
		if playerRowTwo == 0:
			maxRow += 1
			playerRowTwo = maxRow
			table.cell(row = playerRowTwo, column = 1).value = player_two
			table.cell(row = playerRowTwo, column = 11).value = 'E'
			table.cell(row = playerRowTwo, column = 3).value = 1
			table.cell(row = playerRowTwo, column = 4).value = 1
			table.cell(row = playerRowTwo, column = 6).value = 1
			table.cell(row = playerRowTwo, column = 9).value = steps_two
		if win:
			if playing_char == 'O':
				try:
					table.cell(row = playerRowOne, column = 2).value += 1
					table.cell(row = playerRowOne, column = 10).value = 1
					table.cell(row = playerRowOne, column = 10).value = (int(table.cell(row = playerRowOne, column = 10).value) + int(300 / steps_one))
				except:
					table.cell(row = playerRowOne, column = 2).value = 1
					table.cell(row = playerRowOne, column = 10).value = 1
					table.cell(row = playerRowOne, column = 10).value = int(300 / steps_one)
				temp = table.cell(row = playerRowOne, column = 10).value
				if temp >= 10000:
					table.cell(row = playerRowOne, column = 11).value = 'X'
				elif temp >= 8000:
					table.cell(row = playerRowOne, column = 11).value = 'U'
				elif temp >= 6000:
					table.cell(row = playerRowOne, column = 11).value = 'A'
				elif temp >= 4000:
					table.cell(row = playerRowOne, column = 11).value = 'B'
				elif temp >= 2000:
					table.cell(row = playerRowOne, column = 11).value = 'C'
				elif temp >= 800:
					table.cell(row = playerRowOne, column = 11).value = 'D'
			else:
				try:
					table.cell(row = playerRowTwo, column = 2).value += 1
					table.cell(row = playerRowTwo, column = 10).value = 1
					table.cell(row = playerRowTwo, column = 10).value = (int(table.cell(row = playerRowTwo, column = 10).value) + int(300 / steps_two))
				except:
					table.cell(row = playerRowTwo, column = 2).value = 1
					table.cell(row = playerRowTwo, column = 10).value = 1
					table.cell(row = playerRowTwo, column = 10).value = int(300 / steps_two)
				temp = table.cell(row = playerRowTwo, column = 10).value
				if temp >= 10000:
					table.cell(row = playerRowTwo, column = 11).value = 'X'
				elif temp >= 8000:
					table.cell(row = playerRowTwo, column = 11).value = 'U'
				elif temp >= 6000:
					table.cell(row = playerRowTwo, column = 11).value = 'A'
				elif temp >= 4000:
					table.cell(row = playerRowTwo, column = 11).value = 'B'
				elif temp >= 2000:
					table.cell(row = playerRowTwo, column = 11).value = 'C'
				elif temp >= 800:
					table.cell(row = playerRowTwo, column = 11).value = 'D'
		try:
			table.cell(row = playerRowOne, column = 8).value = (table.cell(row = playerRowOne, column = 2).value)/(table.cell(row = playerRowOne, column = 3).value)*100
		except:
			table.cell(row = playerRowOne, column = 8).value = 0
		try:
			table.cell(row = playerRowTwo, column = 8).value = (table.cell(row = playerRowTwo, column = 2).value)/(table.cell(row = playerRowTwo, column = 3).value)*100
		except:
			table.cell(row = playerRowTwo, column = 8).value = 0
	elif mode == 2:
		for i in range(2, maxRow+1):
				temp = table.cell(row = i, column = 1).value
				if  temp == player_one:
					table.cell(row = i, column = 3).value += 1
					if player_char == 'X':
						try:
							table.cell(row = i, column = 4).value += 1
						except:
							table.cell(row = i, column = 4).value = 1
					else:
						try:
							table.cell(row = i, column = 5).value += 1
						except:
							table.cell(row = i, column = 5).value = 1
					try:
						table.cell(row = i, column = 7).value += 1
					except:
						table.cell(row = i,column = 7).value = 1
					table.cell(row = i, column = 9).value = (steps_one + table.cell(row = i, column = 9).value)/2
					playerRowOne = i
		if playerRowOne == 0:
			maxRow += 1
			playerRowOne = maxRow
			table.cell(row = playerRowOne, column = 1).value = player_one
			table.cell(row = playerRowOne, column = 11).value = 'E'
			table.cell(row = playerRowOne, column = 3).value = 1
			if player_char == 'X':
				table.cell(row = playerRowOne, column = 4).value = 1
			elif player_char == 'O':
				table.cell(row = playerRowOne, column = 5).value = 1
			table.cell(row = playerRowOne, column = 7).value = 1
			table.cell(row = playerRowOne, column = 9).value = steps_one
		if win and player_char == playing_char:
			try:
				table.cell(row = playerRowOne, column = 2).value += 1
			except:
				table.cell(row = playerRowOne, column = 2).value = 1
			try:
				table.cell(row = playerRowOne, column = 10).value = (int(table.cell(row = playerRowOne, column = 10).value) + int(300 / steps_one))
			except:
				table.cell(row = playerRowOne, column = 10).value = int(300 / steps_one)
			temp = table.cell(row = playerRowOne, column = 10).value
			if temp >= 10000:
				table.cell(row = playerRowOne, column = 11).value = 'X'
			elif temp >= 8000:
				table.cell(row = playerRowOne, column = 11).value = 'U'
			elif temp >= 6000:
				table.cell(row = playerRowOne, column = 11).value = 'A'
			elif temp >= 4000:
				table.cell(row = playerRowOne, column = 11).value = 'B'
			elif temp >= 2000:
				table.cell(row = playerRowOne, column = 11).value = 'C'
			elif temp >= 800:
				table.cell(row = playerRowOne, column = 11).value = 'D'
		try:
			table.cell(row = playerRowOne, column = 8).value = (table.cell(row = playerRowOne, column = 2).value)/(table.cell(row = playerRowOne, column = 3).value)*100
		except:
			table.cell(row = playerRowOne, column = 8).value = 0
	file.save(filename = 'ScoreTable.xlsx')
	file.close()

#Draw the board grid line
def draw_board():
	global line
	global board
	board = turtle.Screen()
	line = turtle.Turtle()
	line.speed(5)
	line.color("Black")
	line.width("5")
	line.penup()
	line.goto(-150,-150)
	line.pendown()
	for i in range(4):
		line.forward(300)
		line.left(90)
	line.penup()
	line.goto(-150,-50)
	line.pendown()
	line.forward(300)
	line.penup()
	line.goto(-150,50)
	line.pendown()
	line.forward(300)

	line.left(90)
	line.penup()
	line.goto(-50,-150)
	line.pendown()
	line.forward(300)
	line.penup()
	line.goto(50,-150)
	line.pendown()
	line.forward(300)
	line.right(90)
	board.update()

def draw_fig(x, y):
	global playing_char
	global cal_board
	global line
	global board
	global player_char
	global mode
	
	if (x > -150 and x < -50) and (y > 50 and y < 150):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[0][0] = playing_char	
		if playing_char == 'X':
			line.penup()
			line.goto(-149,149)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(-51,149)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(-100,50)
			line.pendown()
			line.circle(50)
			 
	elif (x > -50 and x < 50) and (y > 50 and y < 150):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[0][1] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(-49,149)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(49,149)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(0,50)
			line.pendown()
			line.circle(50)
	elif (x > 50 and x < 150) and (y > 50 and y < 150):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[0][2] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(51,149)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(149,149)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(100,50)
			line.pendown()
			line.circle(50)
	elif (x > -150 and x < -50) and (y > -50 and y < 50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[1][0] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(-149,49)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(-51,49)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(-100,-50)
			line.pendown()
			line.circle(50)
	elif (x > -50 and x < 50) and (y > -50 and y < 50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[1][1] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(-49,49)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(49,49)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(0,-50)
			line.pendown()
			line.circle(50)
	elif (x > 50 and x < 150) and (y > -50 and y < 50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[1][2] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(51,49)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(149,49)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(100,-50)
			line.pendown()
			line.circle(50)
	elif (x > -150 and x < -50) and (y > -150 and y < -50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[2][0] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(-149,-51)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(-51,-51)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(-100,-150)
			line.pendown()
			line.circle(50)
	elif (x > -50 and x < 50) and (y > -150 and y < -50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[2][1] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(-49,-51)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(49,-51)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(0,-150)
			line.pendown()
			line.circle(50)
	elif (x > 50 and x < 150) and (y > -150 and y < -50):
		if not(mode == 2 and (playing_char != player_char)):
			cal_board[2][2] = playing_char
		if playing_char == 'X':
			line.penup()
			line.goto(51,-51)
			line.pendown()
			line.right(45)
			line.forward(140)
			line.penup()
			line.goto(149,-51)
			line.pendown()
			line.right(90)
			line.forward(140)
			line.left(135)
		else:
			line.penup()
			line.goto(100,-150)
			line.pendown()
			line.circle(50)
	board.update()
	check()

#Check whether the player has win or not
def check():
	global cal_board
	global draw
	global win
	global playing_char

	#check win
	for i in range(3):
		x = 0
		for j in range(3):
			if cal_board[i][j] == playing_char:
				x += 1
			if x == 3:
				win = True
	for i in range(3):
		x = 0
		for j in range(3):
			if cal_board[j][i] == playing_char:
				x += 1
			if x == 3:
				win = True
	if ((cal_board[0][0] == playing_char) and (cal_board[1][1] == playing_char)) and (cal_board[2][2] == playing_char):
		win = True
	elif ((cal_board[0][2] == playing_char) and (cal_board[1][1] == playing_char)) and (cal_board[2][0] == playing_char):
		win = True
	#check draw
	x = 0
	for i in range(3):
		for j in range(3):
			if cal_board[i][j] == '/':
				x += 1	
	if x == 0:
		draw = True
	if not (win or draw):
		if playing_char == 'O':
			playing_char = 'X'
		else:
			playing_char = 'O'
def find_xcor(x):
	if x == 0:
		return(100)
	elif x == 1:
		return(0)
	else:
		return(-100)

def find_ycor(y):
	if y == 0:
		return(-100)
	elif y == 1:
		return(0)
	else:
		return(100)

def bot():
	global cal_board

	x = -1
	y = -1

	for i in range(3):
		temp = 0
		for j in range(3):
			if (cal_board[i][j] == playing_char):
				temp += 1
			if temp == 2 :
				for f in range(3):
					if cal_board[i][f] == '/':
						x = i
						y = f
	for i in range(3):
		temp = 0
		for j in range(3):
			if (cal_board[j][i] == playing_char):
				temp += 1
			if temp == 2 :
				for f in range(3):
					if cal_board[f][i] == '/':
						x = f
						y = i
	j = 0
	temp = 0
	for i in range(3):
		if (cal_board[j][i] == playing_char):
			temp += 1
		j += 1
		if temp == 2 :
			j = 0
			for f in range(3):
				if cal_board[j][f] == '/':
					x = j
					y = f
				j += 1

	j = 2
	temp = 0
	for i in range(3):
		if (cal_board[i][j] != playing_char) and (cal_board[i][j] != '/'):
			temp += 1
		j -= 1
		if temp == 2 :
			j = 2
			for f in range(3):
				if cal_board[f][j] == '/':
					x = f
					y = j
				j -= 1
	if x == -1:
		j = 0
		f = 0
		#Check rows
		for i in range(3):
			temp = 0
			for j in range(3):
				if (cal_board[i][j] != playing_char) and (cal_board[i][j] != '/'):
					temp += 1
				if temp == 2 :
					for f in range(3):
						if cal_board[i][f] == '/':
							x = i
							y = f
		for i in range(3):
			temp = 0
			for j in range(3):
				if (cal_board[j][i] != playing_char) and (cal_board[j][i] != '/'):
					temp += 1
				if temp == 2 :
					for f in range(3):
						if cal_board[f][i] == '/':
							x = f
							y = i
		j = 0
		temp = 0
		for i in range(3):
			if (cal_board[j][i] != playing_char) and (cal_board[j][i] != '/'):
				temp += 1
			j += 1
		if temp == 2 :
			j = 0
			for f in range(3):
				if cal_board[j][f] == '/':
					x = j
					y = f
				j += 1
		j = 2
		temp = 0
		for i in range(3):
			if (cal_board[i][j] != playing_char) and (cal_board[i][j] != '/'):
				temp += 1
			j -= 1
		if temp == 2 :
			j = 2
			for f in range(3):
				if cal_board[f][j] == '/':
						x = f
						y = j
				j -= 1
	if x == -1:
		x = random.randint(0,2)
		y = random.randint(0,2)
		while(cal_board[x][y] != '/'):
			x = random.randint(0,2)
			y = random.randint(0,2)
	cal_board[x][y] = playing_char
	#print(cal_board)
	xcor = find_xcor(x)
	ycor = find_ycor(y)
	draw_fig(ycor,xcor)	

def checkfull(x,y):
	global full
	global cal_board
	
	if (x > -150 and x < -50) and (y > 50 and y < 150) and (cal_board[0][0] != '/'):
		 full = True
	elif (x > -50 and x < 50) and (y > 50 and y < 150) and (cal_board[0][1] != '/'):
		full = True
	elif (x > 50 and x < 150) and (y > 50 and y < 150) and (cal_board[0][2] != '/'):
		full = True
	elif (x > -150 and x < -50) and (y > -50 and y < 50) and (cal_board[1][0] != '/'):
		full = True
	elif (x > -50 and x < 50) and (y > -50 and y < 50) and (cal_board[1][1] != '/'):
		full = True
	elif (x > 50 and x < 150) and (y > -50 and y < 50) and (cal_board[1][2] != '/'):
		full = True
	elif (x > -150 and x < -50) and (y > -150 and y < -50) and (cal_board[2][0] != '/'):
		full = True
	elif (x > -50 and x < 50) and (y > -150 and y < -50) and (cal_board[2][1] != '/'):
		full = True
	elif (x > 50 and x < 150) and (y > -150 and y < -50) and (cal_board[2][2] != '/'):
		full = True
	else:
		full = False

def select_mode():
	global mode
	global player_one
	global player_two
	global player_char
	global playing_char
	global draw
	global win
	
	print("Game mode 1 : Player vs Player")
	print("Game mode 2 : Player vs Computer")
	print("***Please be reminder that for game mode 2, player must go first, you don't have a choice")
	try:
		print("Game mode 1 or 2 ? ")
		mode = int(input())
		while mode != 1 and mode != 2:
			print("The number is too large! No such game mode! I mean the number 1 or 2")
			mode = int(input())
	except:
		print("what is that? Hey, I mean the number 1 or 2")
		print("See? There is an error now >:(")
		print(" ")
		select_mode()
	if mode == 1:
		player_one = input('The player that play O : ')
		player_two = input('The player that play X : ')
		playing_char = input("Who will go first? X or O ? ")
		while playing_char != 'X' and playing_char != 'O':
			print("Only enter X or O please!")
			playing_char = input('Who will go first? X or O ? ')
	else:
		player_one = input('The player name is (Hey I mean your name !): ')
		player_char = input('Which one do you want? X or O ? ')
		while player_char != 'X' and player_char != 'O':
			print("Only enter X or O please!")
			player_char = input('Which one do you want? X or O ? ')
		playing_char = player_char
	draw = False
	win = False
	print("The first player is " + playing_char)

def main(x,y):
	global mode
	global cal_board
	global playing_char
	global line
	global board
	global player_one
	global player_two
	global steps_one
	global steps_two
	global win
	global draw
	global full
	global player_char
	
	x = int(x)
	y = int(y)
	checkfull(x,y)
	if (x > -150 and x < 150) and (y > -150 and y < 150) and not(full):
		if mode == 2:
			draw_fig(x,y)
			steps_one += 1
			if not (win or draw):
				bot()
		else:
			if playing_char == 'X':
				steps_two += 1
			else:
				steps_one += 1
			draw_fig(x,y)
	if win:
		print(playing_char + " is the winner")
		scoreLogin()
		turtle.clearscreen()
		turtle.bye()
	elif draw:
		print("Hey, it's a tie !")
		scoreLogin()
		turtle.clearscreen()
		turtle.bye()

#Define global variables
global mode
global cal_board
global playing_char
global line
global board
global steps_one
global steps_two
global player_one
global player_two
global win
global draw
global player_char

steps_one = 0
steps_two = 0
cal_board = [['/','/','/'],['/','/','/'],['/','/','/']]
select_mode()
draw_board()
turtle.onscreenclick(main)
turtle.mainloop()